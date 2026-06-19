import openpyxl
import re
from typing import Dict
from sqlalchemy.orm import Session

from ..models.models import Category, Item, Phase, FloorBudget, ImportLog
from ..database import SessionLocal


def parse_number(val) -> float:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("，", "")
    nums = re.findall(r'[\d.]+', s)
    return float(nums[0]) if nums else 0


def import_excel(filepath: str) -> Dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    db: Session = SessionLocal()

    stats = {"categories": 0, "items_created": 0, "items_updated": 0,
             "phases": 0, "floor_budgets": 0}

    try:
        # === Sheet 1: 预算总览 → categories (rows 9-19) ===
        ws = wb["预算总览"]
        for r in range(9, 20):
            name = ws.cell(row=r, column=1).value
            if not name:
                continue
            control_budget = parse_number(ws.cell(row=r, column=2).value)
            ratio = parse_number(ws.cell(row=r, column=3).value)
            timing = str(ws.cell(row=r, column=4).value or "")
            priority = str(ws.cell(row=r, column=5).value or "")

            existing = db.query(Category).filter(Category.name == name).first()
            if existing:
                existing.control_budget = control_budget
                existing.ratio = ratio
                existing.purchase_timing = timing
                existing.priority = priority
            else:
                db.add(Category(name=name, control_budget=control_budget,
                                ratio=ratio, purchase_timing=timing,
                                priority=priority))
            stats["categories"] += 1

        # === Sheet 2: 采购清单 → items (rows 4-55) ===
        ws = wb["采购清单"]
        for r in range(4, 56):
            name = ws.cell(row=r, column=3).value
            if not name:
                continue

            floor_space = str(ws.cell(row=r, column=1).value or "")
            category = str(ws.cell(row=r, column=2).value or "")
            item_name = str(name)
            attr = str(ws.cell(row=r, column=4).value or "")
            phase = str(ws.cell(row=r, column=5).value or "")
            suggestion = str(ws.cell(row=r, column=6).value or "")
            timing = str(ws.cell(row=r, column=7).value or "")
            budget_min = parse_number(ws.cell(row=r, column=8).value)
            budget_max = parse_number(ws.cell(row=r, column=9).value)
            control_budget = parse_number(ws.cell(row=r, column=10).value)
            brand = str(ws.cell(row=r, column=11).value or "")
            priority = str(ws.cell(row=r, column=12).value or "")
            status = str(ws.cell(row=r, column=13).value or "未开始")
            notes = str(ws.cell(row=r, column=14).value or "")

            existing = db.query(Item).filter(Item.item_name == item_name).first()
            if existing:
                for k, v in {
                    "floor_space": floor_space, "category": category,
                    "attr": attr, "phase": phase, "suggestion": suggestion,
                    "timing": timing, "budget_min": budget_min,
                    "budget_max": budget_max, "control_budget": control_budget,
                    "brand_recommendation": brand, "priority": priority,
                    "status": status, "notes": notes
                }.items():
                    setattr(existing, k, v)
                stats["items_updated"] += 1
            else:
                db.add(Item(
                    floor_space=floor_space, category=category,
                    item_name=item_name, attr=attr, phase=phase,
                    suggestion=suggestion, timing=timing,
                    budget_min=budget_min, budget_max=budget_max,
                    control_budget=control_budget,
                    brand_recommendation=brand, priority=priority,
                    status=status, notes=notes
                ))
                stats["items_created"] += 1

        # === Sheet 3: 装修阶段顺序 → phases (rows 4-12) ===
        ws = wb["装修阶段顺序"]
        for r in range(4, 13):
            name = ws.cell(row=r, column=2).value
            if not name:
                continue
            phase_num_raw = ws.cell(row=r, column=1).value
            phase_num = int(parse_number(phase_num_raw)) if phase_num_raw is not None else 0

            db.add(Phase(
                phase_num=phase_num,
                name=str(name),
                month_range=str(ws.cell(row=r, column=3).value or ""),
                core_tasks=str(ws.cell(row=r, column=4).value or ""),
                must_decide=str(ws.cell(row=r, column=5).value or ""),
                dont_buy_early=str(ws.cell(row=r, column=6).value or ""),
                check_points=str(ws.cell(row=r, column=7).value or ""),
                related_categories=str(ws.cell(row=r, column=8).value or ""),
            ))
            stats["phases"] += 1

        # === Sheet 4: 楼层预算 → floor_budgets (rows 4-9) ===
        ws = wb["楼层预算"]
        for r in range(4, 10):
            floor = ws.cell(row=r, column=1).value
            if not floor:
                continue
            db.add(FloorBudget(
                floor=str(floor),
                spaces=str(ws.cell(row=r, column=2).value or ""),
                must_do=str(ws.cell(row=r, column=3).value or ""),
                budget_min=parse_number(ws.cell(row=r, column=4).value),
                budget_max=parse_number(ws.cell(row=r, column=5).value),
                control_budget=parse_number(ws.cell(row=r, column=6).value),
                can_save=str(ws.cell(row=r, column=7).value or ""),
                must_not_save=str(ws.cell(row=r, column=8).value or ""),
            ))
            stats["floor_budgets"] += 1

        # === 导入日志 ===
        import os
        filename = os.path.basename(filepath)
        log = ImportLog(
            filename=filename,
            items_created=stats["items_created"],
            items_updated=stats["items_updated"],
            items_unchanged=0
        )
        db.add(log)
        db.commit()

    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()

    return stats
