"""
Excel导入服务 — 全量替换模式：新版 Excel 是唯一真相源
清空所有数据后重新导入，确保与 Excel 完全一致
"""
import openpyxl
from typing import Dict
from sqlalchemy.orm import Session
from ..models.models import Category, Item, Phase, FloorBudget
from ..logging_config import get_logger


TRUNCATE_ORDER = [Item, Category, Phase, FloorBudget]


class ExcelImporter:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.logger = get_logger("house_design.importer")

    def _safe_float(self, val) -> float:
        if val is None:
            return 0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0

    def _truncate_all(self, db: Session):
        """清空所有数据表（保留 import_logs）"""
        self.logger.info("Truncating all data tables for full reimport")
        for model in TRUNCATE_ORDER:
            count = db.query(model).count()
            db.query(model).delete()
            self.logger.debug(f"Truncated {model.__name__}: {count} rows")
        db.commit()

    def import_all(self, db: Session) -> Dict[str, int]:
        """全量导入：清库 → 逐个 sheet 导入"""
        self.logger.info("Starting full Excel import", extra={"extra": {"file": self.filepath}})
        self._truncate_all(db)
        result = {}
        result.update(self.import_categories(db))
        result.update(self.import_items(db))
        result.update(self.import_phases(db))
        result.update(self.import_floor_budgets(db))
        return result

    # ── Categories ──
    def import_categories(self, db: Session) -> dict:
        ws = self.wb["预算总览"]
        count = 0
        for row in ws.iter_rows(min_row=9, max_row=20, values_only=True):
            name = row[0]
            if not name:
                continue
            db.add(Category(
                name=str(name).strip(),
                control_budget=self._safe_float(row[1]),
                ratio=self._safe_float(row[2]),
                purchase_timing=str(row[3]) if row[3] else "",
                priority=str(row[4]) if row[4] else "",
            ))
            count += 1
        db.commit()
        self.logger.info("Categories imported", extra={"extra": {"count": count}})
        return {"categories_count": count}

    # ── Items ──
    def import_items(self, db: Session) -> dict:
        ws = self.wb["采购清单"]
        count = 0

        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]
            floor_space, category, item_name, attr, phase, suggestion, timing = vals[0:7]
            budget_min = self._safe_float(vals[7])
            budget_max = self._safe_float(vals[8])
            control_budget = self._safe_float(vals[9])
            brand = vals[10] if len(vals) > 10 else ""
            priority = vals[11] if len(vals) > 11 else ""
            status_ = vals[12] if len(vals) > 12 else "未开始"
            notes = vals[13] if len(vals) > 13 else ""

            if not item_name or item_name in ("采购项", ""):
                continue

            db.add(Item(
                floor_space=floor_space, category=category,
                item_name=item_name, attr=attr, phase=phase,
                suggestion=suggestion, timing=timing,
                budget_min=budget_min, budget_max=budget_max,
                control_budget=control_budget,
                brand_recommendation=brand, priority=priority,
                status=status_ if status_ else "未开始", notes=notes,
            ))
            count += 1
        db.commit()
        self.logger.info("Items imported", extra={"extra": {"count": count}})
        return {"items_count": count}

    # ── Phases ──
    def import_phases(self, db: Session) -> dict:
        ws = self.wb["装修阶段顺序"]
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]
            if not vals[1]:
                continue
            num, name, month, tasks, must, dont, checks, related = vals[0:8]
            db.add(Phase(
                phase_num=int(num) if num.isdigit() else 0,
                name=name, month_range=month,
                core_tasks=tasks, must_decide=must,
                dont_buy_early=dont, check_points=checks,
                related_categories=related,
            ))
            count += 1
        db.commit()
        self.logger.info("Phases imported", extra={"extra": {"count": count}})
        return {"phases_count": count}

    # ── Floor Budgets ──
    def import_floor_budgets(self, db: Session) -> dict:
        self.logger.info("Importing floor budgets...")
        ws = self.wb["楼层预算"]
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]
            floor = vals[0]
            if not floor or floor in ("楼层", "合计"):
                continue
            spaces, must = vals[1], vals[2]
            can_save, must_not = vals[6] if len(vals) > 6 else "", vals[7] if len(vals) > 7 else ""
            db.add(FloorBudget(
                floor=floor, spaces=spaces, must_do=must,
                budget_min=self._safe_float(vals[3]),
                budget_max=self._safe_float(vals[4]),
                control_budget=self._safe_float(vals[5]),
                can_save=can_save, must_not_save=must_not,
            ))
            count += 1
        db.commit()
        return {"floors_count": count}
