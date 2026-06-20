"""
Excel 导出服务 — 将数据库中的状态/实际花费回填到本地 Excel
保持原始 Excel 结构，追加新列，实现双向联动
"""
import openpyxl
from copy import copy
from sqlalchemy.orm import Session
from ..models.models import Item
from .excel_store import get_latest_excel_path, init_store


class ExcelExporter:
    def __init__(self, db: Session):
        self.db = db
        self.items_map = {}
        for item in db.query(Item).all():
            self.items_map[item.item_name] = item

    def export(self, output_path: str = None) -> str:
        """导出 Excel，回填状态/实际花费/供应商，返回文件路径"""
        template = get_latest_excel_path()
        if not template:
            raise FileNotFoundError(
                "未找到本地 Excel 备份。请先通过「导入数据」上传 Excel 文件。"
            )

        wb = openpyxl.load_workbook(template)
        self._update_procurement_sheet(wb)

        if output_path:
            wb.save(output_path)
            return output_path
        else:
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                wb.save(tmp.name)
                return tmp.name

    def _update_procurement_sheet(self, wb):
        """更新采购清单 sheet：回填状态 + 追加实际花费/已支付/供应商列"""
        if "采购清单" not in wb.sheetnames:
            return

        ws = wb["采购清单"]
        header_row = 3
        new_headers = ["实际花费", "已支付", "供应商"]

        # ── 智能查找/追加列 ──
        # 扫描标题行，找到现有列范围和新列位置
        existing_headers = {}
        last_col = 0
        for col in range(1, ws.max_column + 20):  # +20 留余量
            val = ws.cell(row=header_row, column=col).value
            if val is not None:
                existing_headers[str(val).strip()] = col
                last_col = col
            elif col > ws.max_column + 5:
                break  # 连续空列超过5列则停止

        # 检查哪些新列名已存在
        col_map = {}
        next_col = last_col
        for h in new_headers:
            if h in existing_headers:
                col_map[h] = existing_headers[h]  # 复用已有列
            else:
                next_col += 1
                col_map[h] = next_col
                # 添加表头
                cell = ws.cell(row=header_row, column=next_col, value=h)
                if last_col > 0:
                    src_cell = ws.cell(row=header_row, column=last_col)
                    if src_cell.font:
                        cell.font = copy(src_cell.font)
                    if src_cell.fill:
                        cell.fill = copy(src_cell.fill)
                    if src_cell.alignment:
                        cell.alignment = copy(src_cell.alignment)

        # ── 遍历数据行，回填 ──
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            if len(row) < 3:
                continue
            item_name_cell = row[2]  # C 列 = 采购项
            status_cell = row[12] if len(row) > 12 else None  # M 列 = 状态

            item_name = str(item_name_cell.value).strip() if item_name_cell.value else ""
            if not item_name or item_name in ("采购项", ""):
                continue

            db_item = self.items_map.get(item_name)
            if not db_item:
                continue

            row_num = item_name_cell.row

            # 回填状态（M 列 = 第 13 列）
            if status_cell and db_item.status:
                status_cell.value = db_item.status

            # 回填实际花费 / 已支付 / 供应商（按智能定位的列号）
            ws.cell(row=row_num, column=col_map["实际花费"], value=db_item.actual_cost or 0)
            ws.cell(row=row_num, column=col_map["已支付"], value=db_item.actual_paid or 0)
            ws.cell(row=row_num, column=col_map["供应商"], value=db_item.supplier or "")
