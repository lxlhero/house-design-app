"""
Excel 本地存储服务
- 导入时保存 Excel 到本地作为备份真理源
- DB 变更后自动回写 Excel 保持同步
"""
import os
import shutil
from datetime import datetime
import openpyxl
from copy import copy
from pathlib import Path
from sqlalchemy.orm import Session
from ..models.models import Item

# 本地存储目录（相对于项目 backend 目录）
DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
HISTORY_DIR = DATA_DIR / "history"
LOCAL_EXCEL_NAME = "别墅装修_最新.xlsx"
LOCAL_EXCEL_PATH = DATA_DIR / LOCAL_EXCEL_NAME


def init_store():
    """确保 data 目录存在"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)


def save_uploaded_excel(filepath: str) -> str:
    """保存用户上传的 Excel，同时创建时间戳快照"""
    init_store()
    dest = str(LOCAL_EXCEL_PATH)

    # 旧当前文件 → 时间戳快照
    if Path(dest).exists():
        mtime = os.path.getmtime(dest)
        ts = datetime.fromtimestamp(mtime).strftime("%Y%m%d_%H%M%S")
        snapshot_name = f"别墅装修_{ts}.xlsx"
        shutil.copy2(dest, str(HISTORY_DIR / snapshot_name))

    # 新文件覆盖当前
    shutil.copy2(filepath, dest)
    return dest


def list_excel_history() -> list[dict]:
    """列出历史 Excel 快照（按时间倒序）"""
    init_store()
    result = []
    if not HISTORY_DIR.exists():
        return result
    for f in sorted(HISTORY_DIR.glob("别墅装修_*.xlsx"), reverse=True):
        stat = f.stat()
        result.append({
            "filename": f.name,
            "size": stat.st_size,
            "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })
    return result


def get_history_filepath(filename: str) -> str | None:
    """获取历史 Excel 文件的绝对路径（安全检查）"""
    safe_name = os.path.basename(filename)
    if not safe_name.startswith("别墅装修_") or not safe_name.endswith(".xlsx"):
        return None
    path = HISTORY_DIR / safe_name
    return str(path) if path.exists() else None


def get_latest_excel_path() -> str | None:
    """获取最新 Excel 的路径（本地备份）"""
    if LOCAL_EXCEL_PATH.exists():
        return str(LOCAL_EXCEL_PATH)
    return None


def _find_col_map(ws, header_row: int, new_headers: list[str]) -> dict:
    """智能查找/追加列映射"""
    existing = {}
    last_col = 0
    for col in range(1, ws.max_column + 20):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            existing[str(val).strip()] = col
            last_col = col
        elif col > ws.max_column + 5:
            break

    col_map = {}
    next_col = last_col
    for h in new_headers:
        if h in existing:
            col_map[h] = existing[h]
        else:
            next_col += 1
            col_map[h] = next_col
            cell = ws.cell(row=header_row, column=next_col, value=h)
            if last_col > 0:
                src = ws.cell(row=header_row, column=last_col)
                if src.font: cell.font = copy(src.font)
                if src.fill: cell.fill = copy(src.fill)
                if src.alignment: cell.alignment = copy(src.alignment)
    return col_map


def sync_db_to_excel(db: Session) -> str | None:
    """
    将数据库中物品的状态/实际花费/供应商回写到本地 Excel
    返回写入的文件路径，如果没有 Excel 则返回 None
    """
    path = get_latest_excel_path()
    if not path:
        return None

    # 构建 DB 物品映射
    items_map = {}
    for item in db.query(Item).all():
        items_map[item.item_name] = item

    wb = openpyxl.load_workbook(path)

    if "采购清单" not in wb.sheetnames:
        return path

    ws = wb["采购清单"]
    header_row = 3
    col_map = _find_col_map(ws, header_row, ["实际花费", "已支付", "供应商"])

    # 遍历数据行
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        if len(row) < 3:
            continue
        item_name_cell = row[2]
        status_cell = row[12] if len(row) > 12 else None

        item_name = str(item_name_cell.value).strip() if item_name_cell.value else ""
        if not item_name or item_name in ("采购项", ""):
            continue

        db_item = items_map.get(item_name)
        if not db_item:
            continue

        row_num = item_name_cell.row

        # 回填状态
        if status_cell and db_item.status:
            status_cell.value = db_item.status

        # 回填实际花费 / 已支付 / 供应商
        ws.cell(row=row_num, column=col_map["实际花费"], value=db_item.actual_cost or 0)
        ws.cell(row=row_num, column=col_map["已支付"], value=db_item.actual_paid or 0)
        ws.cell(row=row_num, column=col_map["供应商"], value=db_item.supplier or "")

    wb.save(path)
    return path
